from config.settings import settings
from django.core import mail
from django.http import HttpResponse
import openpyxl
from typing import Any, List
import pandas as pd
from teachers.models import Teacher
from students.models import Student
from ratings import models as ratings_models
from subjects import models as subjects_models
from base import constants
from dateutil.relativedelta import relativedelta
from datetime import datetime


def send_email(subject, body):
    success = True
    try:
        mail.send_mail(subject=subject,
                       message=subject,
                       html_message=body,
                       from_email=settings.EMAIL_HOST_USER,
                       recipient_list=[settings.DEFAULT_TO_EMAIL],
                       fail_silently=False
                       )
    except Exception as e:
        print(f"Error: El email no pudo ser enviado!!")
        success = False

    return success


def generate_excel(file_name: str, wb) -> HttpResponse:
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{file_name}.xlsx"'
    wb.save(response)
    return response


def get_copy_template(object_type: str):
    if object_type == constants.TEACHERS:
        wb = openpyxl.load_workbook(f"{settings.APPS_DIR}/fixtures/TEACHERS/plantilla_profesores.xlsx")
        file_name = "Formato registro masivo profesores"
    elif object_type == constants.STUDENTS:
        wb = openpyxl.load_workbook(f"{settings.APPS_DIR}/fixtures/STUDENTS/plantilla_estudiantes.xlsx")
        file_name = "Formato registro masivo estudiantes"
    else:
        wb = openpyxl.load_workbook(f"{settings.APPS_DIR}/fixtures/RATINGS/plantilla_calificaciones.xlsx")
        file_name = "Formato registro masivo calificaciones"

    return generate_excel(file_name=file_name, wb=wb)


def generate_dataframe_records(file: Any, object_type: str, file_type: str):
    success = True
    try:
        if file_type in [constants.XLS, constants.XLSX]:
            io_record = pd.ExcelFile(path_or_buffer=file)
            if file_type == constants.XLSX:
                dataframe = pd.read_excel(io=io_record, engine="openpyxl")
            else:
                dataframe = pd.read_excel(io=io_record)  # for xls files
        else:
            dataframe = pd.read_csv(filepath_or_buffer=file, engine='c')

    except Exception as e:
        print(str(e))
        success = False
    else:
        if object_type == constants.TEACHERS:
            try:
                dataframe = dataframe.rename(
                    columns={
                        "DOCUMENTO": "document_number",
                        "TIPO DOCUMENTO": "document_type",
                        "NOMBRES": "first_name",
                        "APELLIDOS": "last_name",
                        "FECHA NACIMIENTO": "date_of_birth",
                        "GENERO": "gender",
                        "TELEFONO": "phone",
                        "EMAIL": "email",
                        "DIRECCION": "address"
                    }
                )
                dataframe["date_of_birth"] = pd.to_datetime(
                    dataframe["date_of_birth"], errors="coerce"
                )
                dataframe["document_type"] = dataframe["document_type"].str.upper()
                dataframe["gender"] = dataframe["gender"].str.upper()
            except Exception as e:
                print(str(e))
                success = False

        elif object_type == constants.STUDENTS:
            try:
                dataframe = dataframe.rename(
                    columns={
                        "DOCUMENTO": "document_number",
                        "TIPO DOCUMENTO": "document_type",
                        "NOMBRES": "first_name",
                        "APELLIDOS": "last_name",
                        "FECHA NACIMIENTO": "date_of_birth",
                        "GENERO": "gender",
                        "TELEFONO": "phone",
                    }
                )
                dataframe["date_of_birth"] = pd.to_datetime(
                    dataframe["date_of_birth"], errors="coerce"
                )
                dataframe["document_type"] = dataframe["document_type"].str.upper()
                dataframe["gender"] = dataframe["gender"].str.upper()
            except Exception as e:
                print(str(e))
                success = False
        else:
            try:
                dataframe = dataframe.rename(
                    columns={
                        "ACTIVIDAD": "activity",
                        "DOCUMENTO ESTUDIANTE": "student",
                        "DOCUMENTO PROFESOR": "teacher",
                        "CODIGO CURSO": "subject",
                        "CALIFICACION": "rating_number",
                        "OBSERVACION": "observations",
                    }
                )
            except Exception as e:
                print(str(e))
                success = False

        if success:
            return dataframe.to_dict(orient="records")
        return success


def save_records(records: List[dict], object_type: str) -> dict:

    success_amount = 0
    errors_amount = 0

    if object_type == constants.TEACHERS:
        object_name = "Profesor(es)"
        for item in records:
            if str(item["phone"])[0:3] != '+57':
                phone = '+57' + str(item["phone"])
            else:
                phone = item["phone"]

            # validating age
            if relativedelta(datetime.now(), item["date_of_birth"]).years < 18:
                errors_amount += 1
                continue

            try:
                Teacher.objects.create(
                    document_number=item["document_number"],
                    document_type=item["document_type"],
                    first_name=item["first_name"],
                    last_name=item["last_name"],
                    date_of_birth=item["date_of_birth"],
                    gender=item["gender"],
                    phone=phone,
                    email=item["email"],
                    address=item["address"]
                )
                success_amount += 1
            except Exception as e:
                print(str(e))
                errors_amount += 1
                continue

    elif object_type == constants.STUDENTS:
        object_name = "Estudiante(s)"
        for item in records:
            if str(item["phone"])[0:3] != '+57':
                phone = '+57' + str(item["phone"])
            else:
                phone = item["phone"]

            try:
                Student.objects.create(
                    document_number=item["document_number"],
                    document_type=item["document_type"],
                    first_name=item["first_name"],
                    last_name=item["last_name"],
                    date_of_birth=item["date_of_birth"],
                    gender=item["gender"],
                    phone=phone
                )
                success_amount += 1
            except Exception as e:
                print(str(e))
                errors_amount += 1
                continue
    else:
        object_name = "Calificación(es)"
        for item in records:

            # check if active student exist
            try:
                student = Student.objects.get(document_number=item["student"], status=1)
            except Student.DoesNotExist:
                errors_amount += 1
                continue

            # check if active teacher exist
            try:
                teacher = Teacher.objects.get(document_number=item["teacher"], status=1)
            except Teacher.DoesNotExist:
                errors_amount += 1
                continue

            # check if active subject exist
            try:
                subject = subjects_models.Subject.objects.get(code=item["subject"], status=1)
            except subjects_models.Subject.DoesNotExist:
                errors_amount += 1
                continue

            # check if teacher is subject's teacher
            if teacher.id != subject.teacher_id:
                errors_amount += 1
                continue

            # check if student is in subject's students
            if student not in subject.students.all():
                errors_amount += 1
                continue

            try:
                ratings_models.Rating.objects.create(
                    activity=item["activity"],
                    student=student,
                    teacher=teacher,
                    subject=subject,
                    rating_number=item["rating_number"],
                    observations=item["observations"],
                )
                success_amount += 1
            except Exception as e:
                print(str(e))
                errors_amount += 1
                continue

    return {"object_name": object_name, "success_amount": success_amount, "errors_amount": errors_amount}
